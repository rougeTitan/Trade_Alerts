"""
Trade Alerts — Web Dashboard
Flask-based UI for managing watchlist, price targets, and monitoring.
"""

import os
import csv
import io
import json
import tempfile
import threading
import time
from datetime import datetime

import pytz
from flask import Flask, render_template, render_template_string, jsonify, request
from flask_cors import CORS
from werkzeug.utils import secure_filename

from excel_manager import read_watchlist, update_current_prices, create_watchlist_template
from price_fetcher import fetch_prices, fetch_earnings_dates
from monitor import PriceMonitor
from notifier import Notifier

# Run from DATA_DIR when set (e.g. /tmp on AWS Lambda), else the app directory.
_data_dir = os.environ.get("DATA_DIR")
os.chdir(_data_dir if _data_dir else os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True
CORS(app)


@app.after_request
def add_no_cache_headers(response):
    """Prevent browser from caching responses — always serve fresh content."""
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# ---------------------------------------------------------------------------
# Globals for background monitor thread
# ---------------------------------------------------------------------------
monitor_thread = None
monitor_running = False
monitor_lock = threading.Lock()


def _load_config():
    with open("config.json", "r") as f:
        return json.load(f)


def _load_sectors():
    if os.path.exists("sectors.json"):
        with open("sectors.json", "r") as f:
            return json.load(f)
    return {}


def _save_sectors(data):
    with open("sectors.json", "w") as f:
        json.dump(data, f, indent=4)


# ---------------------------------------------------------------------------
# Helper: read / write targets via openpyxl directly (avoids full regenerate)
# ---------------------------------------------------------------------------
import openpyxl


def _ensure_watchlist():
    """Create watchlist.xlsx from sectors.json if it doesn't exist."""
    config = _load_config()
    wl_path = config["files"]["watchlist_excel"]
    if not os.path.exists(wl_path):
        sectors = _load_sectors()
        if sectors:
            create_watchlist_template(sectors, wl_path)


def _set_targets(ticker: str, sector: str, targets: list):
    """
    Write price targets for *ticker* into the watchlist Excel.

    targets: list of {"price": float, "direction": str} (max 3)
    """
    config = _load_config()
    wl_path = config["files"]["watchlist_excel"]
    _ensure_watchlist()

    wb = openpyxl.load_workbook(wl_path)
    saved = False

    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            cell_val = ws[f"A{row}"].value
            if cell_val and str(cell_val).strip().upper() == ticker.upper():
                # Clear existing targets
                for col in ("D", "E", "F", "G", "H", "I"):
                    ws[f"{col}{row}"] = None

                # Write new targets (up to 3)
                col_pairs = [("D", "E"), ("F", "G"), ("H", "I")]
                for i, t in enumerate(targets[:3]):
                    if not t:
                        continue
                    price_col, dir_col = col_pairs[i]
                    ws[f"{price_col}{row}"] = t["price"]
                    ws[f"{dir_col}{row}"] = t.get("direction", "BOTH")
                saved = True
                break
        if saved:
            break

    wb.save(wl_path)
    wb.close()


def _add_ticker_to_excel(ticker: str, sector: str):
    """Append a new ticker row to the appropriate sector sheet."""
    config = _load_config()
    wl_path = config["files"]["watchlist_excel"]
    _ensure_watchlist()

    wb = openpyxl.load_workbook(wl_path)
    if sector not in wb.sheetnames:
        # Create new sheet with headers
        ws = wb.create_sheet(sector)
        headers = ["Ticker", "Company Name", "Current Price",
                    "Price Target 1", "Direction 1",
                    "Price Target 2", "Direction 2",
                    "Price Target 3", "Direction 3", "Notes"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        next_row = 2
    else:
        ws = wb[sector]
        next_row = ws.max_row + 1

    ws[f"A{next_row}"] = ticker.upper()
    wb.save(wl_path)
    wb.close()


def _remove_ticker_from_excel(ticker: str, sector: str):
    """Remove a ticker row from the sector sheet."""
    config = _load_config()
    wl_path = config["files"]["watchlist_excel"]
    if not os.path.exists(wl_path):
        return

    wb = openpyxl.load_workbook(wl_path)
    if sector in wb.sheetnames:
        ws = wb[sector]
        for row in range(2, ws.max_row + 1):
            cell_val = ws[f"A{row}"].value
            if cell_val and str(cell_val).strip().upper() == ticker.upper():
                ws.delete_rows(row)
                break
    wb.save(wl_path)
    wb.close()


# ---------------------------------------------------------------------------
# Background monitor
# ---------------------------------------------------------------------------
def _monitor_loop():
    """Run the price monitor in a background thread."""
    global monitor_running
    import time

    config = _load_config()
    tz = pytz.timezone(config["monitoring"]["timezone"])
    interval = config["monitoring"]["check_interval_seconds"]

    mon = PriceMonitor()
    notifier = Notifier()
    last_day = None
    check_count = 0

    print("🔄 Background monitoring started")

    while monitor_running:
        now = datetime.now(tz)
        today = now.date()
        if last_day != today:
            mon.reset_alerts()
            last_day = today
            check_count = 0

        try:
            # Use cached prices for efficiency (prices are updated by user refreshes)
            # This reduces API calls and improves cloud cost-efficiency
            alerts = mon.check_prices(use_cached=True)
            if alerts:
                print(f"🚨 {len(alerts)} alert(s) detected - sending notifications...")
                notifier.send_alert(alerts)
            check_count += 1
        except Exception as e:
            print(f"❌ Monitor error: {e}")

        # Sleep in small chunks so we can stop quickly
        for _ in range(interval):
            if not monitor_running:
                break
            time.sleep(1)


# ---------------------------------------------------------------------------
# Routes — Pages
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


# ---------------------------------------------------------------------------
# Routes — API
# ---------------------------------------------------------------------------
@app.get("/api/sectors")
def api_get_sectors():
    return jsonify(_load_sectors())


@app.post("/api/sectors/stock")
def api_add_stock():
    data = request.get_json(force=True)
    sector = data.get("sector", "").strip()
    ticker = data.get("ticker", "").strip().upper()
    if not sector or not ticker:
        return jsonify({"error": "sector and ticker required"}), 400

    sectors = _load_sectors()
    if sector not in sectors:
        sectors[sector] = []
    if ticker in sectors[sector]:
        return jsonify({"error": "ticker already exists in sector"}), 409
    sectors[sector].append(ticker)
    _save_sectors(sectors)

    # Also add to Excel
    _add_ticker_to_excel(ticker, sector)
    return jsonify({"ok": True})


def _remove_stock(sector: str, ticker: str):
    sector = (sector or "").strip()
    ticker = (ticker or "").strip().upper()
    sectors = _load_sectors()
    if sector in sectors and ticker in sectors[sector]:
        sectors[sector].remove(ticker)
        if not sectors[sector]:
            del sectors[sector]
        _save_sectors(sectors)
        _remove_ticker_from_excel(ticker, sector)


@app.delete("/api/sectors/stock")
def api_remove_stock():
    """Body-based route (kept for backward compatibility)."""
    data = request.get_json(force=True)
    _remove_stock(data.get("sector", ""), data.get("ticker", ""))
    return jsonify({"ok": True})


@app.delete("/api/sectors/<sector>/stock/<ticker>")
def api_remove_stock_path(sector, ticker):
    """Path-param route. DELETE requests with a body are unreliable through
    CloudFront/OAC, so the client uses this instead."""
    _remove_stock(sector, ticker)
    return jsonify({"ok": True})


@app.post("/api/sectors/add")
def api_add_sector():
    data = request.get_json(force=True)
    sector = data.get("sector", "").strip()
    if not sector:
        return jsonify({"error": "sector name required"}), 400
    sectors = _load_sectors()
    if sector in sectors:
        return jsonify({"error": "sector already exists"}), 409
    sectors[sector] = []
    _save_sectors(sectors)
    return jsonify({"ok": True})


@app.delete("/api/sectors/<sector_name>")
def api_delete_sector(sector_name):
    sectors = _load_sectors()
    if sector_name in sectors:
        del sectors[sector_name]
        _save_sectors(sectors)
    return jsonify({"ok": True})


@app.get("/api/watchlist")
def api_get_watchlist():
    """Return full watchlist with targets, grouped by sector."""
    _ensure_watchlist()
    config = _load_config()
    wl_path = config["files"]["watchlist_excel"]

    try:
        items = read_watchlist(wl_path)
    except FileNotFoundError:
        items = []

    # Also read current prices from the Excel
    prices = {}
    if os.path.exists(wl_path):
        wb = openpyxl.load_workbook(wl_path, data_only=True)
        for sn in wb.sheetnames:
            if sn == "Summary":
                continue
            ws = wb[sn]
            for row in range(2, ws.max_row + 1):
                t = ws[f"A{row}"].value
                p = ws[f"C{row}"].value
                if t:
                    try:
                        prices[str(t).strip().upper()] = float(p) if p else None
                    except (ValueError, TypeError):
                        prices[str(t).strip().upper()] = None
        wb.close()

    # Fetch earnings dates for all tickers
    sectors = _load_sectors()
    all_tickers = [t for tl in sectors.values() for t in tl]
    earnings = fetch_earnings_dates(all_tickers) if all_tickers else {}

    # Merge into a sector-grouped structure
    result = {}
    target_map = {item["ticker"]: item for item in items}

    for sector, tickers in sectors.items():
        result[sector] = []
        for tk in tickers:
            info = target_map.get(tk, {})
            result[sector].append({
                "ticker": tk,
                "current_price": prices.get(tk),
                "targets": info.get("targets", []),
                "earnings_date": earnings.get(tk),
            })
    return jsonify(result)


@app.post("/api/targets")
def api_set_targets():
    data = request.get_json(force=True)
    ticker = data.get("ticker", "").strip().upper()
    sector = data.get("sector", "").strip()
    targets = data.get("targets", [])

    if not ticker:
        return jsonify({"error": "ticker required"}), 400

    # Validate targets
    clean = []
    for t in targets:
        if t is None:
            clean.append(None)
            continue
        try:
            p = float(t["price"])
            d = str(t.get("direction", "BOTH")).upper()
            if d not in ("ABOVE", "BELOW", "BOTH"):
                d = "BOTH"
            clean.append({"price": p, "direction": d})
        except (KeyError, ValueError, TypeError):
            continue

    _set_targets(ticker, sector, clean)
    return jsonify({"ok": True, "saved": len(clean)})


@app.get("/api/prices/refresh")
def api_refresh_prices():
    """Fetch live prices for all tickers and update Excel, then return updated watchlist."""
    sectors = _load_sectors()
    all_tickers = [t for tl in sectors.values() for t in tl]
    if not all_tickers:
        return jsonify({})

    prices = fetch_prices(all_tickers)
    earnings = fetch_earnings_dates(all_tickers)

    config = _load_config()
    wl_path = config["files"]["watchlist_excel"]
    if os.path.exists(wl_path):
        try:
            update_current_prices(wl_path, prices)
        except Exception as e:
            print(f"Error updating prices in Excel: {e}")

    _ensure_watchlist()
    
    try:
        items = read_watchlist(wl_path)
    except FileNotFoundError:
        items = []
    
    target_map = {item["ticker"]: item for item in items}
    
    result = {}
    for sector, tickers in sectors.items():
        result[sector] = []
        for tk in tickers:
            info = target_map.get(tk, {})
            result[sector].append({
                "ticker": tk,
                "current_price": prices.get(tk),
                "targets": info.get("targets", []),
                "earnings_date": earnings.get(tk),
            })
    
    return jsonify(result)


ALERT_MAX_AGE_DAYS = 14


def _alert_age_days(timestamp: str):
    """Age of an alert in days. Returns None if the timestamp can't be parsed
    (so the row is kept rather than silently dropped)."""
    if not timestamp:
        return None
    base = " ".join(str(timestamp).split(" ")[:2])  # drop trailing tz label
    try:
        dt = datetime.strptime(base, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None
    return (datetime.now() - dt).total_seconds() / 86400.0


@app.get("/api/alerts")
def api_get_alerts():
    config = _load_config()
    log_path = config["files"]["alert_log"]
    alerts = []
    fieldnames = None
    pruned = 0
    if os.path.exists(log_path):
        with open(log_path, "r", newline="") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                age = _alert_age_days(row.get("timestamp"))
                if age is not None and age > ALERT_MAX_AGE_DAYS:
                    pruned += 1
                    continue
                alerts.append(row)

    # Durably drop expired alerts from the log so they never reappear.
    if pruned and fieldnames:
        with open(log_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(alerts)

    alerts = list(reversed(alerts))  # newest first
    return jsonify(alerts)


@app.delete("/api/alerts")
def api_clear_alerts():
    config = _load_config()
    log_path = config["files"]["alert_log"]
    if os.path.exists(log_path):
        os.remove(log_path)
    return jsonify({"ok": True, "cleared": True})


@app.delete("/api/alerts/<ticker>/<direction>")
def api_dismiss_alert(ticker, direction):
    config = _load_config()
    log_path = config["files"]["alert_log"]
    if not os.path.exists(log_path):
        return jsonify({"ok": True, "removed": 0})

    rows = []
    fieldnames = None
    removed = 0
    with open(log_path, "r", newline="") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        for row in reader:
            if (row.get("ticker") or "").upper() == ticker.upper() and (row.get("direction") or "").upper() == direction.upper():
                removed += 1
                continue
            rows.append(row)

    if removed:
        with open(log_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    return jsonify({"ok": True, "removed": removed})


@app.get("/api/monitor/status")
def api_monitor_status():
    return jsonify({"running": monitor_running})


@app.post("/api/monitor/start")
def api_monitor_start():
    global monitor_thread, monitor_running
    with monitor_lock:
        if monitor_running:
            return jsonify({"status": "already running"})
        monitor_running = True
        monitor_thread = threading.Thread(target=_monitor_loop, daemon=True)
        monitor_thread.start()
    return jsonify({"status": "started"})


@app.post("/api/monitor/stop")
def api_monitor_stop():
    global monitor_running
    with monitor_lock:
        monitor_running = False
    return jsonify({"status": "stopped"})


@app.post("/api/check-once")
def api_check_once():
    """Run a single price check and return any alerts."""
    mon = PriceMonitor()
    alerts = mon.check_prices()
    return jsonify({"alerts": alerts, "count": len(alerts)})


# ---------------------------------------------------------------------------
# File upload -> S3
# ---------------------------------------------------------------------------
UPLOAD_PAGE = """<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Trade Alerts - Bulk Import</title>
  <style>
    body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; max-width: 640px; margin: 40px auto; padding: 0 20px; background: #0f172a; color: #e2e8f0; }
    h1 { font-size: 24px; margin-bottom: 8px; }
    p, li { color: #94a3b8; margin-bottom: 12px; }
    form { background: #1e293b; padding: 24px; border-radius: 12px; border: 1px solid #334155; }
    input[type="file"] { color: #e2e8f0; margin-bottom: 20px; }
    button { background: #22c55e; color: #fff; border: none; border-radius: 8px; padding: 12px 20px; font-weight: 600; cursor: pointer; }
    a { color: #22c55e; text-decoration: none; }
    code { background: #334155; padding: 2px 6px; border-radius: 4px; color: #e2e8f0; }
    pre { background: #1e293b; padding: 12px; border-radius: 8px; overflow-x: auto; font-size: 13px; color: #e2e8f0; }
    .hint { font-size: 14px; }
  </style>
</head>
<body>
  <h1>Bulk Import</h1>
  <p class="hint">Upload a CSV or the legacy watchlist.xlsx. Each row becomes a stock in your profile.</p>

  <p class="hint">CSV columns (header required):</p>
  <pre>ticker,sector,current_price,target1,dir1,target2,dir2,target3,dir3</pre>
  <p class="hint"><code>dir</code> = ABOVE, BELOW or BOTH. Leave blanks for missing targets or current price.</p>

  <p class="hint">Example:</p>
  <pre>ticker,sector,current_price,target1,dir1,target2,dir2,target3,dir3
AAPL,Technology,150.00,160.00,ABOVE,140.00,BELOW,,
MSFT,Technology,250.00,260.00,ABOVE,,,,</pre>

  <form action="/api/upload" method="post" enctype="multipart/form-data">
    <input type="hidden" name="token" value="{{ token }}" />
    <input type="file" name="file" accept=".csv,.xlsx" required />
    <br />
    <button type="submit">Import</button>
  </form>
</body>
</html>"""


@app.route("/api/upload", methods=["GET", "POST"])
def api_upload():
    token = request.values.get("token", "")
    if request.method == "GET":
        return render_template_string(UPLOAD_PAGE, token=token)

    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    filename = secure_filename(file.filename)
    if not filename:
        return jsonify({"error": "Invalid filename"}), 400

    # Authenticated v2 bulk import -> DynamoDB
    if token:
        if not os.environ.get("DYNAMODB_TABLE"):
            return jsonify({"error": "DynamoDB not configured"}), 503

        try:
            from auth import get_user_from_token
            user = get_user_from_token(token)
            if not user:
                return _result_html("Invalid or expired token", "fail", 401)
            user_id = user["user_id"]
        except Exception as e:
            return _result_html(f"Token check failed: {e}", "fail", 500)

        ext = filename.rsplit(".", 1)[-1].lower()
        if ext not in ("csv", "xlsx"):
            return _result_html("Only .csv and .xlsx are supported for import", "fail", 400)

        try:
            if ext == "csv":
                added, skipped, errors = _bulk_import_csv(file, user_id)
            else:
                added, skipped, errors = _bulk_import_xlsx(file, user_id)
        except Exception as e:
            return _result_html(f"Import failed: {e}", "fail", 500)

        message = f"Imported {added} stocks, skipped {skipped}."
        if errors:
            message += f" Errors ({len(errors)}): {'; '.join(errors[:5])}"
        return _result_html(message, "ok" if not errors else "fail")

    # Legacy unauthenticated upload -> S3 or local
    key = f"uploads/{int(time.time())}_{filename}"
    bucket = os.environ.get("STATE_BUCKET")

    if bucket:
        try:
            import boto3
            s3 = boto3.client("s3")
            s3.upload_fileobj(file, bucket, key)
            return _result_html(f"Stored as s3://{bucket}/{key}", "ok")
        except Exception as e:
            return _result_html(f"Upload failed: {e}", "fail", 500)

    # Local fallback
    upload_dir = os.path.join(_data_dir or os.path.dirname(os.path.abspath(__file__)), "uploads")
    os.makedirs(upload_dir, exist_ok=True)
    path = os.path.join(upload_dir, os.path.basename(key))
    file.save(path)
    return _result_html(f"Saved locally to {path}", "ok")


def _result_html(message, status, code=200):
    color = "#86efac" if status == "ok" else "#fca5a5"
    bg = "#14532d" if status == "ok" else "#7f1d1d"
    icon = "✅" if status == "ok" else "❌"
    return f"""<!DOCTYPE html>
<html><body style="font-family:sans-serif;max-width:640px;margin:40px auto;background:#0f172a;color:#e2e8f0">
  <div style="background:{bg};color:{color};padding:16px;border-radius:8px;margin-bottom:16px">
    <h2 style="margin:0">{icon} {message}</h2>
  </div>
  <p><a href="/api/upload" style="color:#22c55e">Back</a></p>
</body></html>""", code


def _bulk_import_csv(file, user_id):
    import db
    reader = csv.DictReader(io.TextIOWrapper(file, encoding="utf-8-sig"))
    added = 0
    skipped = 0
    errors = []
    for i, row in enumerate(reader, start=2):
        ticker = (row.get("ticker") or "").strip().upper()
        sector = (row.get("sector") or "").strip()
        if not ticker or not sector:
            skipped += 1
            continue
        try:
            db.add_stock(user_id, sector, ticker)
            targets = _parse_csv_targets(row)
            if any(t is not None for t in targets):
                db.set_targets(user_id, sector, ticker, targets)
            current_price = (row.get("current_price") or "").strip()
            if current_price:
                db.set_stock_price(user_id, sector, ticker, float(current_price))
            added += 1
        except Exception as e:
            errors.append(f"row {i}: {e}")
    return added, skipped, errors


def _parse_csv_targets(row):
    targets = []
    for pkey, dkey in [("target1", "dir1"), ("target2", "dir2"), ("target3", "dir3")]:
        price = (row.get(pkey) or "").strip()
        if not price:
            targets.append(None)
            continue
        try:
            p = float(price)
            d = (row.get(dkey) or "BOTH").strip().upper() or "BOTH"
            if d not in ("ABOVE", "BELOW", "BOTH"):
                d = "BOTH"
            targets.append({"price": p, "direction": d})
        except ValueError:
            targets.append(None)
    return targets


def _bulk_import_xlsx(file, user_id):
    import db
    from excel_manager import read_watchlist
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file.save(tmp)
        tmp_path = tmp.name
    try:
        rows = read_watchlist(tmp_path)
    except Exception as e:
        raise RuntimeError(f"Could not read xlsx: {e}")
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
    added = 0
    skipped = 0
    errors = []
    for item in rows:
        ticker = (item.get("ticker") or "").strip().upper()
        sector = (item.get("sector") or "").strip()
        if not ticker or not sector:
            skipped += 1
            continue
        try:
            db.add_stock(user_id, sector, ticker)
            targets = item.get("targets") or [None, None, None]
            if any(t is not None for t in targets):
                db.set_targets(user_id, sector, ticker, targets)
            added += 1
        except Exception as e:
            errors.append(f"{ticker}: {e}")
    return added, skipped, errors


# ---------------------------------------------------------------------------
# Auto-start monitoring
# ---------------------------------------------------------------------------
def _auto_start_monitoring():
    """Automatically start background monitoring when app starts."""
    global monitor_thread, monitor_running
    config = _load_config()
    
    # Only auto-start if not already running
    with monitor_lock:
        if not monitor_running:
            monitor_running = True
            monitor_thread = threading.Thread(target=_monitor_loop, daemon=True)
            monitor_thread.start()
            print("✅ Background alert monitoring auto-started")
            print(f"   Check interval: {config['monitoring']['check_interval_seconds']} seconds")
            print(f"   Email alerts: {'ON' if config['email']['enabled'] else 'OFF'}")
            print(f"   SMS alerts: {'ON' if config['sms']['enabled'] else 'OFF'}")


# ---------------------------------------------------------------------------
# Initialize on import so it also runs under gunicorn (no __main__ there).
# Idempotent + cheap.
_ensure_watchlist()

# Monitor auto-start is env-gated to avoid double-running (e.g. when a separate
# systemd "main.py monitor" process handles monitoring). Default ON for local
# dev and single-worker gunicorn. Set AUTO_START_MONITOR=0 to disable.
if os.environ.get("AUTO_START_MONITOR", "1") == "1":
    _auto_start_monitoring()


# Multi-user v2 API — only loaded when DynamoDB + Cognito are configured.
# Keeps the legacy /api routes alive for local dev without AWS.
if os.environ.get("DYNAMODB_TABLE"):
    from api_v2 import api_v2
    app.register_blueprint(api_v2, url_prefix="/api/v2")


if __name__ == "__main__":
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug, host=host, port=port)
