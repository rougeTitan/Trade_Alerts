"""
Excel Watchlist Generator & Reader
Generates a watchlist Excel template organized by sector for the user to fill in price targets.
Also reads the completed watchlist back for monitoring.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json


def create_watchlist_template(sectors_data: dict, output_file: str = "watchlist.xlsx"):
    """
    Create an Excel watchlist template organized by sector.
    
    Args:
        sectors_data: dict like {
            "Technology": ["AAPL", "MSFT", "GOOGL"],
            "Energy": ["XOM", "CVX"],
            ...
        }
        output_file: path to save the Excel file
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sector_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    ticker_font = Font(name="Calibri", size=11)
    price_font = Font(name="Calibri", size=11, color="006100")
    
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    
    # --- SUMMARY SHEET ---
    summary_ws = wb.create_sheet("Summary")
    summary_ws["A1"] = "STOCK PRICE ALERT WATCHLIST"
    summary_ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="2F5496")
    summary_ws["A3"] = "Instructions:"
    summary_ws["A3"].font = Font(name="Calibri", bold=True, size=12)
    
    instructions = [
        "1. Each sector has its own tab/sheet below.",
        "2. For each stock, fill in up to 3 Price Targets in the green columns.",
        "3. Price Target 1, 2, 3 can be in any order (above or below current price).",
        "4. The system will alert you when the stock price crosses ANY of your targets.",
        "5. Leave a price target blank if you don't need all 3.",
        "6. The 'Direction' column: enter 'ABOVE' or 'BELOW' to specify alert direction.",
        "   - ABOVE = alert when price goes ABOVE the target",
        "   - BELOW = alert when price goes BELOW the target",
        "   - If left blank, alerts trigger in BOTH directions.",
        "7. Save this file and run the monitor script.",
        "",
        "Sectors in this watchlist:"
    ]
    
    for i, line in enumerate(instructions, start=4):
        summary_ws[f"A{i}"] = line
        summary_ws[f"A{i}"].font = Font(name="Calibri", size=11)
    
    row = 4 + len(instructions)
    for sector_name, tickers in sectors_data.items():
        summary_ws[f"A{row}"] = f"  • {sector_name} ({len(tickers)} stocks)"
        summary_ws[f"A{row}"].font = Font(name="Calibri", size=11, bold=True)
        row += 1
    
    summary_ws.column_dimensions["A"].width = 80
    
    # --- SECTOR SHEETS ---
    for sector_name, tickers in sectors_data.items():
        # Truncate sheet name to 31 chars (Excel limit)
        sheet_name = sector_name[:31]
        ws = wb.create_sheet(sheet_name)
        
        # Column headers
        headers = [
            ("A", "Ticker", 14),
            ("B", "Company Name", 30),
            ("C", "Current Price", 16),
            ("D", "Price Target 1", 16),
            ("E", "Direction 1", 14),
            ("F", "Price Target 2", 16),
            ("G", "Direction 2", 14),
            ("H", "Price Target 3", 16),
            ("I", "Direction 3", 14),
            ("J", "Notes", 30),
        ]
        
        for col_letter, header_text, width in headers:
            cell = ws[f"{col_letter}1"]
            cell.value = header_text
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[col_letter].width = width
        
        # Fill in tickers
        for i, ticker in enumerate(tickers, start=2):
            ws[f"A{i}"] = ticker.upper()
            ws[f"A{i}"].font = Font(name="Calibri", size=11, bold=True)
            ws[f"A{i}"].alignment = center_align
            ws[f"A{i}"].border = border
            
            # Company name placeholder
            ws[f"B{i}"].font = ticker_font
            ws[f"B{i}"].border = border
            
            # Current price (will be filled by script)
            ws[f"C{i}"].font = ticker_font
            ws[f"C{i}"].alignment = center_align
            ws[f"C{i}"].border = border
            ws[f"C{i}"].number_format = '$#,##0.00'
            
            # Price targets - highlighted in green
            target_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            direction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            for col in ["D", "F", "H"]:  # Price target columns
                ws[f"{col}{i}"].fill = target_fill
                ws[f"{col}{i}"].font = price_font
                ws[f"{col}{i}"].alignment = center_align
                ws[f"{col}{i}"].border = border
                ws[f"{col}{i}"].number_format = '$#,##0.00'
            
            for col in ["E", "G", "I"]:  # Direction columns
                ws[f"{col}{i}"].fill = direction_fill
                ws[f"{col}{i}"].font = ticker_font
                ws[f"{col}{i}"].alignment = center_align
                ws[f"{col}{i}"].border = border
            
            # Notes
            ws[f"J{i}"].font = ticker_font
            ws[f"J{i}"].border = border
        
        # Freeze top row
        ws.freeze_panes = "A2"
    
    wb.save(output_file)
    print(f"✅ Watchlist template saved to: {output_file}")
    print(f"   Sectors: {len(sectors_data)}")
    print(f"   Total stocks: {sum(len(t) for t in sectors_data.values())}")
    return output_file


def read_watchlist(input_file: str = "watchlist.xlsx") -> list:
    """
    Read the completed watchlist Excel file and return a list of stock alert configs.
    
    Returns:
        List of dicts:
        [
            {
                "sector": "Technology",
                "ticker": "AAPL",
                "targets": [
                    {"price": 150.00, "direction": "ABOVE"},
                    {"price": 140.00, "direction": "BELOW"},
                    {"price": 160.00, "direction": "ABOVE"},
                ]
            },
            ...
        ]
    """
    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Watchlist file not found: {input_file}")
    
    wb = openpyxl.load_workbook(input_file, data_only=True)
    watchlist = []
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        
        ws = wb[sheet_name]
        sector = sheet_name
        
        for row in range(2, ws.max_row + 1):
            ticker = ws[f"A{row}"].value
            if not ticker or str(ticker).strip() == "":
                continue
            
            ticker = str(ticker).strip().upper()
            targets = []
            
            # Read 3 price targets and directions
            target_cols = [("D", "E"), ("F", "G"), ("H", "I")]
            for price_col, dir_col in target_cols:
                price_val = ws[f"{price_col}{row}"].value
                dir_val = ws[f"{dir_col}{row}"].value
                
                if price_val is not None:
                    try:
                        price = float(price_val)
                        direction = str(dir_val).strip().upper() if dir_val else "BOTH"
                        # Normalize common aliases
                        if direction in ("DOWN", "D"):
                            direction = "BELOW"
                        elif direction in ("UP", "U"):
                            direction = "ABOVE"
                        if direction not in ("ABOVE", "BELOW", "BOTH"):
                            direction = "BOTH"
                        targets.append({
                            "price": price,
                            "direction": direction
                        })
                    except (ValueError, TypeError):
                        targets.append(None)
                else:
                    targets.append(None)
            
            if any(targets):
                watchlist.append({
                    "sector": sector,
                    "ticker": ticker,
                    "targets": targets
                })
    
    wb.close()
    return watchlist


def update_current_prices(input_file: str, prices: dict):
    """
    Update the 'Current Price' column in the watchlist with live prices.
    
    Args:
        input_file: path to watchlist Excel
        prices: dict like {"AAPL": 150.00, "MSFT": 380.00, ...}
    """
    wb = openpyxl.load_workbook(input_file)
    
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            ticker = ws[f"A{row}"].value
            if ticker and str(ticker).strip().upper() in prices:
                ws[f"C{row}"] = prices[str(ticker).strip().upper()]
    
    wb.save(input_file)


if __name__ == "__main__":
    # Example usage - this will be replaced with actual data from screenshots
    sample_data = {
        "Technology": ["AAPL", "MSFT", "GOOGL", "NVDA", "META"],
        "Energy": ["XOM", "CVX", "COP"],
        "Healthcare": ["JNJ", "PFE", "UNH"],
    }
    create_watchlist_template(sample_data, "watchlist_SAMPLE.xlsx")
